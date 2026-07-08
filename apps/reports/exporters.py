"""
AKHU AFIVS — Report Exporters
PDF (ReportLab), Excel (openpyxl), CSV exports
"""
import csv
import io
import logging
from datetime import datetime
from typing import List, Dict, Any

from django.utils import timezone
from django.utils.translation import gettext as _
from django.http import HttpResponse

logger = logging.getLogger(__name__)


# ─── HELPERS ────────────────────────────────────────────────────────────────

def _get_report_title(report_type: str) -> str:
    titles = {
        'verified': _('Verified Applicants Report'),
        'failed': _('Failed Verifications Report'),
        'exam_attendance': _('Examination Attendance Report'),
        'daily_activity': _('Daily Verification Activity Report'),
        'supervisor_activity': _('Supervisor Activity Logs Report'),
        'security_incidents': _('Security Incident Reports'),
        'attendance': _('Examination Attendance Report'),
        'pre_registered': _('Allowed Candidates Report'),
        'audit_logs': _('System Audit Logs Report'),
        'daily_stats': _('Daily Verification Stats Report'),
    }
    return titles.get(report_type, _('Report'))


def _format_date(dt) -> str:
    if dt is None:
        return '-'
    if hasattr(dt, 'strftime'):
        return dt.strftime('%Y-%m-%d %H:%M')
    return str(dt)


# ─── CSV EXPORT ─────────────────────────────────────────────────────────────

def export_csv(data: List[Dict], headers: List[str], filename: str) -> HttpResponse:
    """Export data as CSV response."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    response.write('\ufeff')  # BOM for Excel UTF-8 compatibility

    writer = csv.DictWriter(response, fieldnames=headers)
    writer.writeheader()
    for row in data:
        filtered = {k: v for k, v in row.items() if k in headers}
        writer.writerow(filtered)

    return response


# ─── EXCEL EXPORT ───────────────────────────────────────────────────────────

def export_excel(data: List[Dict], headers: List[str], title: str, filename: str) -> HttpResponse:
    """Export data as Excel (.xlsx) response."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error('openpyxl not installed')
        return HttpResponse('openpyxl not installed', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit

    # ── Styles ──
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
    title_font = Font(name='Arial', bold=True, size=14, color='0A1628')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # ── Title Row ──
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = center_align

    # ── Generated At Row ──
    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    gen_cell = ws['A2']
    gen_cell.value = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")} | AKHU Face Verification System'
    gen_cell.alignment = center_align
    gen_cell.font = Font(italic=True, size=9, color='666666')

    # ── Header Row ──
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── Data Rows ──
    for row_idx, row in enumerate(data, start=4):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else '')
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            # Alternate row color
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')

    # ── Column Widths ──
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    ws.row_dimensions[3].height = 25

    # ── Freeze Header ──
    ws.freeze_panes = 'A4'

    # ── Save ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


# ─── PDF EXPORT ─────────────────────────────────────────────────────────────

def export_pdf(data: List[Dict], headers: List[str], title: str, filename: str) -> HttpResponse:
    """Export data as PDF response using ReportLab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        logger.error('reportlab not installed')
        return HttpResponse('reportlab not installed', status=500)

    buffer = io.BytesIO()
    page_size = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    AKHU_BLUE = colors.HexColor('#0A1628')
    AKHU_GOLD = colors.HexColor('#F59E0B')

    # Custom styles
    title_style = ParagraphStyle(
        'AKHUTitle', parent=styles['Title'],
        fontSize=16, textColor=AKHU_BLUE, spaceAfter=6,
        alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        'AKHUSub', parent=styles['Normal'],
        fontSize=9, textColor=colors.grey,
        alignment=TA_CENTER, spaceAfter=12,
    )

    # ── Build flowables ──
    elements = []

    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f'AKHU Face Identity Verification System | Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}',
        sub_style,
    ))
    elements.append(HRFlowable(width='100%', thickness=2, color=AKHU_BLUE))
    elements.append(Spacer(1, 0.5 * cm))

    # Summary stats
    elements.append(Paragraph(
        f'Total Records: <b>{len(data)}</b>',
        ParagraphStyle('stats', parent=styles['Normal'], fontSize=10),
    ))
    elements.append(Spacer(1, 0.3 * cm))

    # ── Table ──
    table_data = [headers]
    for row in data:
        table_data.append([str(row.get(h, ''))[:50] for h in headers])

    col_count = len(headers)
    page_w = page_size[0] - 3 * cm
    col_w = page_w / col_count

    table = Table(table_data, colWidths=[col_w] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), AKHU_BLUE),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4FF')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


# ─── REPORT DATA BUILDERS ────────────────────────────────────────────────────

def get_verified_applicants_data() -> tuple:
    """Return (headers, data) for verified applicants report."""
    from apps.verification.models import FaceProfile
    from apps.accounts.models import ApplicantProfile

    profiles = ApplicantProfile.objects.select_related(
        'user',
        'qr_code',
    ).filter(
        user__verification_sessions__face_profile__status='verified',
    ).distinct()

    headers = ['#', 'Admission ID', 'Full Name', 'Passport Number',
               'Email', 'Phone', 'Match %', 'QR Token', 'Verified At']
    data = []
    for i, p in enumerate(profiles, 1):
        face = FaceProfile.objects.filter(
            session__user=p.user, status='verified'
        ).order_by('-created_at').first()
        qr = getattr(p, 'qr_code', None)
        data.append({
            '#': i,
            'Admission ID': p.admission_id,
            'Full Name': p.full_name,
            'Passport Number': p.passport_number,
            'Email': p.email,
            'Phone': p.phone_number,
            'Match %': f'{face.match_percentage:.1f}%' if face else '-',
            'QR Token': qr.token if qr else '-',
            'Verified At': _format_date(face.created_at if face else None),
        })
    return headers, data


def get_failed_verifications_data() -> tuple:
    """Return (headers, data) for failed verifications report."""
    from apps.verification.models import FaceProfile

    profiles = FaceProfile.objects.filter(
        status='rejected'
    ).select_related('session__user').order_by('-created_at')

    headers = ['#', 'Session ID', 'Match %', 'Liveness', 'Anti-Spoof', 'Attempted At', 'IP']
    data = []
    for i, fp in enumerate(profiles, 1):
        data.append({
            '#': i,
            'Session ID': str(fp.session.id)[:12],
            'Match %': f'{fp.match_percentage:.1f}%' if fp.match_percentage else '-',
            'Liveness': 'Passed' if fp.liveness_passed else 'Failed',
            'Anti-Spoof': fp.anti_spoof_result or '-',
            'Attempted At': _format_date(fp.created_at),
            'IP': fp.session.ip_address or '-',
        })
    return headers, data


def get_supervisor_activity_data() -> tuple:
    """Return (headers, data) for supervisor activity report."""
    from apps.verification.models import VerificationLog

    logs = VerificationLog.objects.filter(
        verification_type='exam_day'
    ).select_related('supervisor', 'applicant_profile').order_by('-verified_at')

    headers = ['#', 'Supervisor', 'Applicant', 'Admission ID', 'Result', 'Score', 'Date/Time', 'IP']
    data = []
    for i, log in enumerate(logs, 1):
        data.append({
            '#': i,
            'Supervisor': log.supervisor.username if log.supervisor else '-',
            'Applicant': log.applicant_profile.full_name if log.applicant_profile else '-',
            'Admission ID': log.applicant_profile.admission_id if log.applicant_profile else '-',
            'Result': log.result,
            'Score': f'{log.score:.1f}%' if log.score else '-',
            'Date/Time': _format_date(log.verified_at),
            'IP': log.ip_address or '-',
        })
    return headers, data


def get_attendance_report_data() -> tuple:
    """Return (headers, data) for exam attendance report."""
    from apps.admin_panel.views import _attendance_rows
    rows = _attendance_rows()
    
    headers = ['#', 'Applicant ID', 'Full Name', 'Passport', 'Program', 'Region', 'Status', 'Registered At', 'Verification Date', 'Attendance Time', 'Supervisor']
    data = []
    for i, r in enumerate(rows, 1):
        data.append({
            '#': i,
            'Applicant ID': r['applicant_id'] or '-',
            'Full Name': r['full_name'] or '-',
            'Passport': r['passport_number'] or '-',
            'Program': r['program'] or '-',
            'Region': r['region'] or '-',
            'Status': r['status'].upper() if r['status'] else '-',
            'Registered At': _format_date(r['registered_at']),
            'Verification Date': _format_date(r['verification_date']),
            'Attendance Time': _format_date(r['attendance_at']),
            'Supervisor': r['supervisor'] or '-',
        })
    return headers, data


def get_pre_registered_data() -> tuple:
    """Return (headers, data) for allowed candidates list."""
    from apps.accounts.models import PreRegisteredApplicant
    candidates = PreRegisteredApplicant.objects.all().order_by('-created_at')
    
    headers = ['#', 'Applicant ID', 'Passport Number', 'Full Name', 'Program', 'Region', 'Uploaded At']
    data = []
    for i, c in enumerate(candidates, 1):
        data.append({
            '#': i,
            'Applicant ID': c.applicant_id or '-',
            'Passport Number': c.passport_number or '-',
            'Full Name': f'{c.surname} {c.given_name} {c.middle_name}'.strip(),
            'Program': c.program or '-',
            'Region': c.region or '-',
            'Uploaded At': _format_date(c.created_at),
        })
    return headers, data


def get_audit_logs_data() -> tuple:
    """Return (headers, data) for system audit logs."""
    from apps.audit.models import AuditLog
    logs = AuditLog.objects.all().order_by('-timestamp')
    
    headers = ['#', 'Timestamp', 'User', 'Role', 'Category', 'Action', 'Status', 'IP Address']
    data = []
    for i, log in enumerate(logs, 1):
        data.append({
            '#': i,
            'Timestamp': _format_date(log.timestamp),
            'User': log.username_snapshot or 'System',
            'Role': log.user_role_snapshot or '-',
            'Category': log.category or '-',
            'Action': log.action or '-',
            'Status': 'Success' if log.success else 'Failed',
            'IP Address': log.ip_address or '-',
        })
    return headers, data


def get_daily_stats_data() -> tuple:
    """Return (headers, data) for daily verification stats report."""
    from apps.verification.models import VerificationLog
    from django.db.models import Count, Q
    from django.db.models.functions import TruncDate
    
    stats = VerificationLog.objects.annotate(date=TruncDate('verified_at')).values('date').annotate(
        total=Count('id'),
        verified=Count('id', filter=Q(result='verified')),
        review=Count('id', filter=Q(result='review_required')),
        rejected=Count('id', filter=Q(result='rejected'))
    ).order_by('-date')
    
    headers = ['#', 'Date', 'Total Attempts', 'Verified', 'Review Required', 'Rejected']
    data = []
    for i, s in enumerate(stats, 1):
        data.append({
            '#': i,
            'Date': str(s['date']) if s['date'] else '-',
            'Total Attempts': s['total'],
            'Verified': s['verified'],
            'Review Required': s['review'],
            'Rejected': s['rejected'],
        })
    return headers, data
