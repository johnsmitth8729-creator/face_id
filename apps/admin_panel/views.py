"""
AKHU AFIVS — Admin Panel Views
Full administrative dashboard with 10 management modules.
"""
import logging
import hmac
from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.views.generic import TemplateView
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.contrib.auth import authenticate
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg
from django.db.models.functions import TruncDate

from apps.accounts.models import CustomUser, ApplicantProfile, SupervisorAccount, PreRegisteredApplicant, ExamVenueConfig
from apps.verification.models import (
    VerificationSession, FaceProfile, VerificationLog, VerificationStatus
)
from apps.audit.models import AuditLog
from apps.qr_module.models import QRCode

logger = logging.getLogger(__name__)

ADMIN_SESSION_KEY = 'admin_authenticated'


def _excel_image_map(sheet, image_col: int | None) -> dict[int, bytes]:
    """Return embedded Excel images keyed by 1-based row number."""
    if not image_col:
        return {}
    images: dict[int, bytes] = {}
    for image in getattr(sheet, '_images', []):
        marker = getattr(getattr(image, 'anchor', None), '_from', None)
        if not marker or marker.col + 1 != image_col:
            continue
        try:
            images[marker.row + 1] = image._data()
        except Exception as exc:
            logger.warning("Could not read embedded Excel image at row %s: %s", marker.row + 1, exc)
    return images


def _decode_cell_image(value) -> bytes | None:
    """Decode base64/data-url image content if supplied directly in a cell."""
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        import base64
        if text.lower().startswith('data:image') and ',' in text:
            text = text.split(',', 1)[1]
        return base64.b64decode(text, validate=True)
    except Exception:
        return None


def _validate_image_bytes(img_bytes: bytes, row_label: str) -> tuple[bytes | None, str]:
    """Validate upload size/type and normalize image bytes to JPEG."""
    if not img_bytes:
        return None, f'{row_label}: Passport Image is required'
    if len(img_bytes) > settings.MAX_UPLOAD_SIZE:
        return None, f'{row_label}: Passport Image exceeds file size limit'

    from PIL import Image, UnidentifiedImageError
    import io

    try:
        image = Image.open(io.BytesIO(img_bytes))
        if image.format not in {'JPEG', 'JPG', 'PNG'}:
            return None, f'{row_label}: Passport Image must be JPEG or PNG'
        image = image.convert('RGB')
    except (UnidentifiedImageError, OSError):
        return None, f'{row_label}: Passport Image cannot be decoded'

    output = io.BytesIO()
    image.save(output, format='JPEG', quality=92)
    return output.getvalue(), ''


def _passport_embedding_from_bytes(img_bytes: bytes, row_label: str) -> tuple[list[float] | None, str]:
    """Generate a strict single-face passport embedding."""
    from PIL import Image
    import io
    from apps.face_engine.engine import get_face_engine

    try:
        image = Image.open(io.BytesIO(img_bytes)).convert('RGB')
    except Exception:
        return None, f'{row_label}: Passport Image cannot be decoded'

    result = get_face_engine().extract_face_and_embedding(image, require_single=True)
    if not result.get('success'):
        face_count = result.get('face_count', 0)
        if face_count == 0:
            return None, f'{row_label}: Passport Image contains zero faces'
        if face_count > 1:
            return None, f'{row_label}: Passport Image contains multiple faces'
        return None, f"{row_label}: Passport embedding generation failed ({result.get('error') or 'unknown error'})"

    embedding = result.get('embedding')
    if embedding is None:
        return None, f'{row_label}: Passport embedding generation failed'
    values = embedding.tolist()
    if len(values) != 512:
        return None, f'{row_label}: Passport embedding must contain 512 values'
    return values, ''


def _admin_required(view_func):
    """Decorator for function-based admin views."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.session.get(ADMIN_SESSION_KEY):
            return redirect(settings.ADMIN_PANEL_LOGIN_URL)
        return view_func(request, *args, **kwargs)
    return wrapper


def admin_required_class(cls):
    """Class-based decorator for admin auth. Also allows exam_staff role for allowed views."""
    original_dispatch = cls.dispatch

    def new_dispatch(self, request, *args, **kwargs):
        allowed_classes = [
            'PermitsManagementView', 
            'DownloadAllPermitsZipView', 
            'AttendanceManagementView', 
            'AttendanceExportExcelView',
            'StatisticsView',
            'StatisticsExportExcelView'
        ]
        class_name = self.__class__.__name__
        
        is_exam_staff_authorized = (
            request.user.is_authenticated 
            and request.user.role == 'exam_staff' 
            and class_name in allowed_classes
        )
        
        if not request.session.get(ADMIN_SESSION_KEY) and not is_exam_staff_authorized:
            return redirect(settings.ADMIN_PANEL_LOGIN_URL)
        return original_dispatch(self, request, *args, **kwargs)

    cls.dispatch = new_dispatch
    return cls


# ─── AUTH ────────────────────────────────────────────────────────────────────

class AdminLoginView(View):
    template_name = 'admin_panel/login.html'

    def get(self, request):
        if request.session.get(ADMIN_SESSION_KEY):
            return redirect('admin_panel:dashboard')
        return render(request, self.template_name, {'page_title': _('Admin Login')})

    def post(self, request):
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        # Validate against .env credentials using constant-time comparison
        # to prevent timing attacks that could enumerate valid credentials.
        env_username = settings.ADMIN_USERNAME
        env_password = settings.ADMIN_PASSWORD

        username_match = hmac.compare_digest(username, env_username)
        password_match = hmac.compare_digest(password, env_password)

        if username_match and password_match:
            request.session[ADMIN_SESSION_KEY] = True
            request.session['admin_username'] = username
            logger.info(f'Admin login: {username} from {request.META.get("REMOTE_ADDR")}')

            AuditLog.objects.create(
                username_snapshot=username,
                user_role_snapshot='admin',
                category='auth',
                action='Admin login',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:200],
                success=True,
            )
            return redirect('admin_panel:dashboard')

        AuditLog.objects.create(
            username_snapshot=username,
            user_role_snapshot='admin',
            category='security',
            action='Failed admin login attempt',
            ip_address=request.META.get('REMOTE_ADDR'),
            success=False,
            error_message='Invalid credentials',
        )
        messages.error(request, _('Invalid administrator credentials'))
        return render(request, self.template_name, {'page_title': _('Admin Login')})


class AdminLogoutView(View):
    def get(self, request):
        request.session.pop(ADMIN_SESSION_KEY, None)
        return redirect('admin_panel:login')


# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@admin_required_class
class AdminDashboardView(TemplateView):
    template_name = 'admin_panel/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.now().date()

        total_applicants = ApplicantProfile.objects.count()
        verified = FaceProfile.objects.filter(status='verified').count()
        pending = FaceProfile.objects.filter(status='review_required').count()
        rejected = FaceProfile.objects.filter(status='rejected').count()

        today_logs = VerificationLog.objects.filter(verified_at__date=today)
        today_total = today_logs.count()

        recent_verifications = VerificationLog.objects.select_related(
            'applicant_profile', 'supervisor'
        ).order_by('-verified_at')[:10]

        # Chart data: last 7 days
        from datetime import timedelta
        days_data = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            count = VerificationLog.objects.filter(verified_at__date=d).count()
            days_data.append({'date': str(d), 'count': count})

        ctx.update({
            'page_title': _('Admin Dashboard'),
            'total_applicants': total_applicants,
            'verified_count': verified,
            'pending_count': pending,
            'rejected_count': rejected,
            'today_total': today_total,
            'recent_verifications': recent_verifications,
            'chart_data': days_data,
            'supervisor_count': SupervisorAccount.objects.filter(is_active=True).count(),
            'qr_count': QRCode.objects.filter(status='active').count(),
        })
        return ctx


# ─── APPLICANT MANAGEMENT ────────────────────────────────────────────────────

@admin_required_class
class ApplicantListView(View):
    template_name = 'admin_panel/applicants.html'

    def get(self, request):
        query = request.GET.get('q', '').strip()

        # 1. Standard/Verified applicants
        applicants = ApplicantProfile.objects.select_related('user').prefetch_related(
            'user__verification_sessions__face_profile',
            'qr_code',
        ).order_by('-created_at')

        # 2. Pre-registered/Allowed candidates list
        pre_registered = PreRegisteredApplicant.objects.all().order_by('-created_at')

        if query:
            applicants = applicants.filter(
                Q(admission_id__icontains=query) |
                Q(applicant_id__icontains=query) |
                Q(passport_number__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(program__icontains=query) |
                Q(selected_region__icontains=query)
            )
            pre_registered = pre_registered.filter(
                Q(passport_number__icontains=query) |
                Q(applicant_id__icontains=query) |
                Q(given_name__icontains=query) |
                Q(surname__icontains=query) |
                Q(middle_name__icontains=query) |
                Q(program__icontains=query) |
                Q(region__icontains=query)
            )

        paginator = Paginator(applicants, 25)
        page = paginator.get_page(request.GET.get('page'))

        pre_paginator = Paginator(pre_registered, 25)
        pre_page = pre_paginator.get_page(request.GET.get('pre_page'))

        return render(request, self.template_name, {
            'page_title': _('Applicant Management'),
            'applicants': page,
            'pre_registered': pre_page,
            'query': query,
        })

    def post(self, request):
        action = request.POST.get('action')

        if action == 'edit_applicant':
            profile_id = request.POST.get('profile_id')
            profile = get_object_or_404(ApplicantProfile.objects.select_related('user'), id=profile_id)

            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            passport_number = request.POST.get('passport_number', '').strip().upper()
            admission_id = request.POST.get('admission_id', '').strip()

            if not first_name or not last_name or not passport_number or not admission_id:
                messages.error(request, _('First name, last name, passport number, and admission ID are required.'))
                return redirect('admin_panel:applicants')

            if ApplicantProfile.objects.exclude(id=profile.id).filter(passport_number=passport_number).exists():
                messages.error(request, _('Another applicant already uses this passport number.'))
                return redirect('admin_panel:applicants')

            if ApplicantProfile.objects.exclude(id=profile.id).filter(admission_id=admission_id).exists():
                messages.error(request, _('Another applicant already uses this admission ID.'))
                return redirect('admin_panel:applicants')

            from django.utils.dateparse import parse_date, parse_datetime
            from django.utils.timezone import make_aware, is_naive

            date_of_birth = parse_date(request.POST.get('date_of_birth', '').strip()) if request.POST.get('date_of_birth') else None
            exam_date = request.POST.get('exam_date', '').strip()
            arrival_time = request.POST.get('arrival_time', '').strip()

            profile.first_name = first_name
            profile.last_name = last_name
            profile.middle_name = request.POST.get('middle_name', '').strip()
            profile.date_of_birth = date_of_birth
            profile.gender = request.POST.get('gender', '').strip()
            profile.phone_number = request.POST.get('phone_number', '').strip()
            profile.email = request.POST.get('email', '').strip()
            profile.passport_number = passport_number
            profile.admission_id = admission_id
            profile.applicant_id = request.POST.get('applicant_id', '').strip() or None
            profile.program = request.POST.get('program', '').strip()
            profile.selected_region = request.POST.get('selected_region', '').strip()
            profile.exam_venue = request.POST.get('exam_venue', '').strip()
            profile.exam_date = exam_date
            profile.arrival_time = arrival_time
            profile.is_locked = bool(request.POST.get('is_locked'))
            profile.save()

            profile.user.email = profile.email
            profile.user.is_active = bool(request.POST.get('user_is_active'))
            profile.user.save(update_fields=['email', 'is_active'])

            AuditLog.objects.create(
                username_snapshot=request.session.get('admin_username', 'admin'),
                user_role_snapshot='admin',
                category='data',
                action=f'Edited applicant: {profile.full_name} ({profile.passport_number})',
                ip_address=request.META.get('REMOTE_ADDR'),
                success=True,
            )
            messages.success(request, _('Applicant information updated successfully.'))

        elif action == 'delete':
            profile_id = request.POST.get('profile_id')
            if profile_id:
                try:
                    profile = ApplicantProfile.objects.get(id=profile_id)
                    name = profile.full_name
                    # Remove biometric media files before CASCADE-deleting the user.
                    try:
                        from apps.verification.cleanup import cleanup_all_incomplete_for_user
                        _admin_username = request.session.get('admin_username', 'admin')
                        cleanup_all_incomplete_for_user(
                            profile.user,
                            reason='admin_reset',
                            performed_by=f'admin:{_admin_username}',
                        )
                    except Exception as _ce:
                        import logging as _lg
                        _lg.getLogger(__name__).warning(
                            "Biometric cleanup before admin delete failed for %s: %s", name, _ce
                        )
                    profile.user.delete()  # cascades
                    AuditLog.objects.create(
                        username_snapshot=request.session.get('admin_username', 'admin'),
                        user_role_snapshot='admin',
                        category='data',
                        action=f'Deleted applicant: {name}',
                        ip_address=request.META.get('REMOTE_ADDR'),
                    )
                    messages.success(request, _(f'Applicant {name} deleted'))
                except ApplicantProfile.DoesNotExist:
                    messages.error(request, _('Applicant not found'))

        elif action == 'bulk_delete':
            selected_ids = request.POST.getlist('selected_ids')
            if selected_ids:
                deleted_count = 0
                names = []
                for p_id in selected_ids:
                    try:
                        profile = ApplicantProfile.objects.get(id=p_id)
                        name = profile.full_name
                        names.append(name)
                        try:
                            from apps.verification.cleanup import cleanup_all_incomplete_for_user
                            _admin_username = request.session.get('admin_username', 'admin')
                            cleanup_all_incomplete_for_user(
                                profile.user,
                                reason='admin_reset',
                                performed_by=f'admin:{_admin_username}',
                            )
                        except Exception as _ce:
                            logger.warning("Biometric cleanup before admin delete failed for %s: %s", name, _ce)
                        profile.user.delete()
                        deleted_count += 1
                    except ApplicantProfile.DoesNotExist:
                        pass
                if deleted_count > 0:
                    AuditLog.objects.create(
                        username_snapshot=request.session.get('admin_username', 'admin'),
                        user_role_snapshot='admin',
                        category='data',
                        action=f'Bulk deleted {deleted_count} applicants: {", ".join(names[:10])}',
                        ip_address=request.META.get('REMOTE_ADDR'),
                    )
                    messages.success(request, _(f'Successfully deleted {deleted_count} applicants.'))
                else:
                    messages.warning(request, _('No applicants were deleted.'))
            else:
                messages.error(request, _('No applicants selected for deletion.'))

        elif action == 'delete_pre':
            pre_id = request.POST.get('pre_id')
            if pre_id:
                try:
                    pre = PreRegisteredApplicant.objects.get(id=pre_id)
                    name = f'{pre.surname} {pre.given_name}'
                    passport = pre.passport_number
                    pre.delete()
                    AuditLog.objects.create(
                        username_snapshot=request.session.get('admin_username', 'admin'),
                        user_role_snapshot='admin',
                        category='data',
                        action=f'Deleted pre-registered candidate: {name} ({passport})',
                        ip_address=request.META.get('REMOTE_ADDR'),
                    )
                    messages.success(request, _(f'Pre-registered candidate {name} deleted'))
                except PreRegisteredApplicant.DoesNotExist:
                    messages.error(request, _('Pre-registered candidate not found'))

        elif action == 'bulk_delete_pre':
            selected_ids = request.POST.getlist('selected_ids')
            if selected_ids:
                deleted_qs = PreRegisteredApplicant.objects.filter(id__in=selected_ids)
                deleted_count = deleted_qs.count()
                passports = list(deleted_qs.values_list('passport_number', flat=True))
                deleted_qs.delete()
                if deleted_count > 0:
                    AuditLog.objects.create(
                        username_snapshot=request.session.get('admin_username', 'admin'),
                        user_role_snapshot='admin',
                        category='data',
                        action=f'Bulk deleted {deleted_count} pre-registered candidates: {", ".join(passports[:10])}',
                        ip_address=request.META.get('REMOTE_ADDR'),
                    )
                    messages.success(request, _(f'Successfully deleted {deleted_count} pre-registered candidates.'))
                else:
                    messages.warning(request, _('No candidates were deleted.'))
            else:
                messages.error(request, _('No candidates selected for deletion.'))

        elif action == 'create_candidate':
            applicant_id = request.POST.get('applicant_id', '').strip()
            surname = request.POST.get('surname', '').strip()
            given_name = request.POST.get('given_name', '').strip()
            middle_name = request.POST.get('middle_name', '').strip()
            program = request.POST.get('program', '').strip()
            passport_number = request.POST.get('passport_number', '').strip().upper()

            if not given_name or not surname or not passport_number:
                messages.error(request, _('Given name, surname, and card/passport number are required'))
                return redirect('admin_panel:applicants')

            if PreRegisteredApplicant.objects.filter(passport_number=passport_number).exists():
                messages.error(request, _('Candidate with this passport number is already registered'))
                return redirect('admin_panel:applicants')

            if applicant_id and PreRegisteredApplicant.objects.filter(applicant_id=applicant_id).exists():
                messages.error(request, _('Candidate with this Applicant ID is already registered'))
                return redirect('admin_panel:applicants')

            from django.db import IntegrityError
            try:
                candidate = PreRegisteredApplicant.objects.create(
                    applicant_id=applicant_id,
                    surname=surname,
                    given_name=given_name,
                    middle_name=middle_name,
                    program=program,
                    region='',
                    passport_number=passport_number,
                    passport_image=None,
                    passport_embedding=None,
                )
            except IntegrityError:
                messages.error(request, _('Database error: Candidate with this Passport or Applicant ID already exists.'))
                return redirect('admin_panel:applicants')

            AuditLog.objects.create(
                username_snapshot=request.session.get('admin_username', 'admin'),
                user_role_snapshot='admin',
                category='data',
                action=f'Created pre-registered candidate: {surname} {given_name} ({passport_number})',
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            messages.success(request, _(f'Candidate {surname} {given_name} registered successfully'))

        elif action == 'import_excel':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, _('Please upload an Excel file'))
                return redirect('admin_panel:applicants')

            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active
                
                # Check headers
                header_row = 1
                for r in range(1, 10):
                    row_vals = [str(sheet.cell(r, c).value).strip().lower() if sheet.cell(r, c).value else '' for c in range(1, sheet.max_column + 1)]
                    if any(h in row_vals for h in ['passport', 'pasport', 'passport number', 'passport_number', 'card / passport number']):
                        header_row = r
                        break
                        
                headers = [str(sheet.cell(header_row, c).value).strip().lower() if sheet.cell(header_row, c).value else '' for c in range(1, sheet.max_column + 1)]
                
                def find_col(names):
                    for idx, h in enumerate(headers):
                        if any(n in h for n in names):
                            return idx + 1
                    return None
                    
                col_applicant_id = find_col(['applicant id', 'applicant_id', 'id'])
                col_surname = find_col(['surname', 'last name', 'lastname', 'familya', 'familiya'])
                col_given_name = find_col(['given name', 'first name', 'firstname', 'ism'])
                col_middle_name = find_col(['middle name', 'sharif', 'middle_name', 'patronymic'])
                col_program = find_col(['program', 'yonalish', 'yo\'nalish', 'mutaxassislik'])
                col_region = find_col(['region', 'hudud', 'viloyat'])
                col_passport = find_col(['card / passport number', 'passport number', 'passport', 'pasport', 'passport_number', 'card'])
                col_passport_image = find_col(['passport image', 'passport photo', 'photo', 'image', 'rasm'])
                
                # Fallbacks if headers not detected
                if not col_applicant_id: col_applicant_id = 1
                if not col_surname: col_surname = 2
                if not col_given_name: col_given_name = 3
                if not col_middle_name: col_middle_name = 4
                if not col_program: col_program = 5
                if not col_passport: col_passport = 6
                embedded_images = _excel_image_map(sheet, col_passport_image)

                imported_count = 0
                errors = []
                
                for r in range(header_row + 1, sheet.max_row + 1):
                    passport_val = sheet.cell(r, col_passport).value if col_passport else None
                    surname_val = str(sheet.cell(r, col_surname).value or '').strip() if col_surname else ''
                    given_name_val = str(sheet.cell(r, col_given_name).value or '').strip() if col_given_name else ''
                    middle_name_val = str(sheet.cell(r, col_middle_name).value or '').strip() if col_middle_name else ''
                    program_val = str(sheet.cell(r, col_program).value or '').strip() if col_program else ''
                    region_val = str(sheet.cell(r, col_region).value or '').strip() if col_region else ''
                    
                    # Skip completely empty rows
                    if not passport_val and not surname_val and not given_name_val:
                        continue

                    if not passport_val:
                        errors.append(f"Row {r}: Card / Passport Number is required")
                        continue
                    
                    passport_val = str(passport_val).strip().upper()

                    if not given_name_val or not surname_val:
                        errors.append(f"Row {r}: Given Name and Surname are required")
                        continue

                    applicant_id_val = str(sheet.cell(r, col_applicant_id).value or '').strip() if col_applicant_id else ''
                    if applicant_id_val.endswith('.0'):
                        applicant_id_val = applicant_id_val[:-2]

                    candidate, created = PreRegisteredApplicant.objects.update_or_create(
                        passport_number=passport_val,
                        defaults={
                            'applicant_id': applicant_id_val,
                            'surname': surname_val,
                            'given_name': given_name_val,
                            'middle_name': middle_name_val,
                            'program': program_val,
                            'region': '',
                            'passport_embedding': None,
                            'passport_image': None,
                        }
                    )
                    imported_count += 1
                
                AuditLog.objects.create(
                    username_snapshot=request.session.get('admin_username', 'admin'),
                    user_role_snapshot='admin',
                    category='data',
                    action=f'Imported {imported_count} candidates via Excel',
                    ip_address=request.META.get('REMOTE_ADDR'),
                )
                
                if imported_count > 0:
                    messages.success(request, _(f'Successfully imported {imported_count} candidates!'))
                if errors:
                    messages.warning(request, _(f'Import completed with warnings: {", ".join(errors[:5])}'))
                    
            except Exception as e:
                messages.error(request, _(f'Failed to parse Excel: {str(e)}'))
        return redirect('admin_panel:applicants')


# ─── PERMITS MANAGEMENT ──────────────────────────────────────────────────────

@admin_required_class
class PermitsManagementView(View):
    template_name = 'admin_panel/permits.html'

    @staticmethod
    def _attach_verified_session(applicants):
        for applicant in applicants:
            verified_session = None
            for session in applicant.user.verification_sessions.all():
                face_profile = getattr(session, 'face_profile', None)
                if face_profile and face_profile.status == VerificationStatus.VERIFIED:
                    verified_session = session
                    break
            applicant.verified_session = verified_session
            applicant.verification_date = (
                getattr(verified_session, 'completed_at', None)
                or getattr(getattr(verified_session, 'face_profile', None), 'created_at', None)
                or getattr(verified_session, 'started_at', None)
            )
        return applicants

    def get(self, request):
        query = request.GET.get('q', '').strip()

        # Get verified applicants who have a locked status or verified biometrics
        applicants = ApplicantProfile.objects.filter(
            user__verification_sessions__face_profile__status='verified'
        ).select_related('user').prefetch_related('user__verification_sessions__face_profile', 'qr_code').distinct().order_by('-created_at')

        if query:
            applicants = applicants.filter(
                Q(applicant_id__icontains=query) |
                Q(passport_number__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(program__icontains=query)
            )

        paginator = Paginator(applicants, 20)
        page = paginator.get_page(request.GET.get('page'))
        self._attach_verified_session(page.object_list)

        return render(request, self.template_name, {
            'page_title': _('Permits Management'),
            'applicants': page,
            'query': query,
        })


@admin_required_class
class DownloadAllPermitsZipView(View):
    def get(self, request):
        import io
        import zipfile

        # Query all verified applicants
        applicants = ApplicantProfile.objects.filter(
            user__verification_sessions__face_profile__status='verified'
        ).select_related('user').prefetch_related('user__verification_sessions__face_profile').distinct()

        if not applicants.exists():
            messages.warning(request, _('No verified applicants found to generate permits.'))
            return redirect('admin_panel:permits')

        # Create in-memory zip
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            from apps.reports.views import generate_confirmation_pdf_stream
            
            for app in applicants:
                # Find the verification session that contains the verified face profile
                session = app.user.verification_sessions.filter(face_profile__status='verified').first()
                if not session:
                    continue
                
                # Generate PDF stream
                pdf_buffer = io.BytesIO()
                success = generate_confirmation_pdf_stream(session, pdf_buffer)
                if success:
                    pdf_data = pdf_buffer.getvalue()
                    pdf_buffer.close()
                    
                    # File name logic: use applicant_id, fallback to passport_number
                    filename = f"{app.applicant_id}.pdf" if app.applicant_id else f"permit_{app.passport_number}.pdf"
                    zip_file.writestr(filename, pdf_data)
                else:
                    pdf_buffer.close()

        zip_buffer.seek(0)
        
        # Log this bulk action in AuditLog
        AuditLog.objects.create(
            username_snapshot=request.session.get('admin_username', 'admin'),
            user_role_snapshot='admin',
            category='data',
            action=f'Bulk downloaded {applicants.count()} permits as ZIP',
            ip_address=request.META.get('REMOTE_ADDR'),
            success=True
        )

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="all_permits.zip"'
        zip_buffer.close()
        return response



# ─── SUPERVISOR MANAGEMENT ───────────────────────────────────────────────────

ATTENDANCE_NOTE = 'Checked-in/Exam Entry Confirmed by Supervisor'


def _attendance_rows(query='', status='', source=''):
    profiles = ApplicantProfile.objects.select_related('user').prefetch_related(
        'user__verification_sessions__face_profile',
        'verification_logs__supervisor',
    )
    pre_registered = PreRegisteredApplicant.objects.all()

    if query:
        profiles = profiles.filter(
            Q(applicant_id__icontains=query) |
            Q(admission_id__icontains=query) |
            Q(passport_number__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(middle_name__icontains=query) |
            Q(program__icontains=query) |
            Q(selected_region__icontains=query)
        )
        pre_registered = pre_registered.filter(
            Q(applicant_id__icontains=query) |
            Q(passport_number__icontains=query) |
            Q(given_name__icontains=query) |
            Q(surname__icontains=query) |
            Q(middle_name__icontains=query) |
            Q(program__icontains=query) |
            Q(region__icontains=query)
        )

    profile_by_passport = {profile.passport_number: profile for profile in profiles}
    rows = []

    def latest_verified_session(profile):
        sessions = sorted(profile.user.verification_sessions.all(), key=lambda item: item.started_at, reverse=True)
        for session in sessions:
            face_profile = getattr(session, 'face_profile', None)
            if face_profile and face_profile.status == VerificationStatus.VERIFIED:
                return session
        return None

    def latest_attendance(profile):
        logs = [
            log for log in profile.verification_logs.all()
            if log.verification_type == 'exam_day' and log.notes == ATTENDANCE_NOTE
        ]
        return sorted(logs, key=lambda item: item.verified_at, reverse=True)[0] if logs else None

    for profile in profiles:
        verified_session = latest_verified_session(profile)
        attendance_log = latest_attendance(profile)
        if attendance_log:
            row_status = 'attended'
        elif verified_session:
            row_status = 'verified_absent'
        else:
            row_status = 'registered'

        rows.append({
            'source': 'profile',
            'status': row_status,
            'applicant_id': profile.applicant_id or profile.admission_id,
            'full_name': profile.full_name,
            'passport_number': profile.passport_number,
            'program': profile.program,
            'region': profile.selected_region,
            'registered_at': profile.created_at,
            'verification_date': (
                getattr(verified_session, 'completed_at', None)
                or getattr(getattr(verified_session, 'face_profile', None), 'created_at', None)
                or getattr(verified_session, 'started_at', None)
            ),
            'attendance_at': getattr(attendance_log, 'verified_at', None),
            'supervisor': getattr(getattr(attendance_log, 'supervisor', None), 'username', ''),
        })

    for pre in pre_registered.exclude(passport_number__in=list(profile_by_passport.keys())):
        rows.append({
            'source': 'allowed',
            'status': 'not_started',
            'applicant_id': pre.applicant_id,
            'full_name': f'{pre.surname} {pre.given_name} {pre.middle_name}'.strip(),
            'passport_number': pre.passport_number,
            'program': pre.program,
            'region': pre.region,
            'registered_at': pre.created_at,
            'verification_date': None,
            'attendance_at': None,
            'supervisor': '',
        })

    if status:
        rows = [row for row in rows if row['status'] == status]
    if source:
        rows = [row for row in rows if row['source'] == source]

    status_order = {'attended': 0, 'verified_absent': 1, 'registered': 2, 'not_started': 3}
    rows.sort(key=lambda row: row.get('attendance_at') or row.get('verification_date') or row.get('registered_at'), reverse=True)
    rows.sort(key=lambda row: status_order.get(row['status'], 9))
    return rows


@admin_required_class
class AttendanceManagementView(View):
    template_name = 'admin_panel/attendance.html'

    def get(self, request):
        query = request.GET.get('q', '').strip()
        status = request.GET.get('status', '').strip()
        source = request.GET.get('source', '').strip()
        rows = _attendance_rows(query=query, status=status, source=source)

        stats = {
            'total': len(rows),
            'attended': sum(1 for row in rows if row['status'] == 'attended'),
            'verified_absent': sum(1 for row in rows if row['status'] == 'verified_absent'),
            'registered': sum(1 for row in rows if row['status'] == 'registered'),
            'not_started': sum(1 for row in rows if row['status'] == 'not_started'),
        }

        paginator = Paginator(rows, 30)
        page = paginator.get_page(request.GET.get('page'))

        return render(request, self.template_name, {
            'page_title': _('Exam Attendance'),
            'attendance_rows': page,
            'query': query,
            'status_filter': status,
            'source_filter': source,
            'stats': stats,
            'status_choices': [
                ('attended', _('Attended')),
                ('verified_absent', _('Verified, Not Attended')),
                ('registered', _('Registered, Not Verified')),
                ('not_started', _('Allowed, Not Started')),
            ],
            'source_choices': [
                ('profile', _('Submitted Profiles')),
                ('allowed', _('Allowed Candidates')),
            ],
        })


@admin_required_class
class AttendanceExportExcelView(View):
    def get(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        rows = _attendance_rows(
            query=request.GET.get('q', '').strip(),
            status=request.GET.get('status', '').strip(),
            source=request.GET.get('source', '').strip(),
        )
        status_labels = {
            'attended': 'Attended',
            'verified_absent': 'Verified, Not Attended',
            'registered': 'Registered, Not Verified',
            'not_started': 'Allowed, Not Started',
        }
        source_labels = {'profile': 'Submitted Profile', 'allowed': 'Allowed Candidate'}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        headers = [
            'Applicant ID', 'Full Name', 'Passport', 'Program', 'Region',
            'Source', 'Status', 'Registered At', 'Verification Date',
            'Attendance Time', 'Supervisor',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        for row in rows:
            ws.append([
                row['applicant_id'] or '',
                row['full_name'] or '',
                row['passport_number'] or '',
                row['program'] or '',
                row['region'] or '',
                source_labels.get(row['source'], row['source']),
                status_labels.get(row['status'], row['status']),
                row['registered_at'].strftime('%Y-%m-%d %H:%M') if row['registered_at'] else '',
                row['verification_date'].strftime('%Y-%m-%d %H:%M') if row['verification_date'] else '',
                row['attendance_at'].strftime('%Y-%m-%d %H:%M') if row['attendance_at'] else '',
                row['supervisor'] or '',
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 38)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="exam_attendance.xlsx"'
        return response


@admin_required_class
class SupervisorManagementView(View):
    template_name = 'admin_panel/supervisors.html'

    def get(self, request):
        from django.db.models import Count, Q
        supervisors = SupervisorAccount.objects.select_related('user', 'created_by').annotate(
            total_scans=Count('user__supervised_verifications', filter=Q(user__supervised_verifications__verification_type='exam_day')),
            verified_scans=Count('user__supervised_verifications', filter=Q(user__supervised_verifications__verification_type='exam_day', user__supervised_verifications__result='verified')),
            failed_scans=Count('user__supervised_verifications', filter=Q(user__supervised_verifications__verification_type='exam_day', user__supervised_verifications__result='rejected')),
        ).order_by('-created_at')

        log_q = request.GET.get('log_q', '').strip()
        activity_logs = VerificationLog.objects.filter(
            verification_type='exam_day',
            supervisor__isnull=False
        ).select_related('applicant_profile', 'supervisor').order_by('-verified_at')

        if log_q:
            activity_logs = activity_logs.filter(
                Q(supervisor__username__icontains=log_q) |
                Q(supervisor__supervisor_account__full_name__icontains=log_q) |
                Q(applicant_profile__first_name__icontains=log_q) |
                Q(applicant_profile__last_name__icontains=log_q) |
                Q(applicant_profile__passport_number__icontains=log_q) |
                Q(applicant_profile__admission_id__icontains=log_q) |
                Q(applicant_profile__applicant_id__icontains=log_q) |
                Q(notes__icontains=log_q)
            )
        else:
            activity_logs = activity_logs[:50]

        from apps.accounts.models import UserRole
        exam_staff_users = CustomUser.objects.filter(role=UserRole.EXAM_STAFF).order_by('-date_joined')

        return render(request, self.template_name, {
            'page_title': _('Supervisor & Staff Management'),
            'supervisors': supervisors,
            'exam_staff_users': exam_staff_users,
            'activity_logs': activity_logs,
            'log_q': log_q,
        })

    def post(self, request):
        from apps.accounts.models import UserRole
        action = request.POST.get('action')

        if action == 'create':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()
            full_name = request.POST.get('full_name', '').strip()

            if not username or not password or not full_name:
                messages.error(request, _('All fields are required'))
                return redirect('admin_panel:supervisors')

            if CustomUser.objects.filter(username=username).exists():
                messages.error(request, _('Username already exists'))
                return redirect('admin_panel:supervisors')

            user = CustomUser.objects.create_supervisor(
                username=username,
                password=password,
            )
            SupervisorAccount.objects.create(
                user=user,
                full_name=full_name,
            )
            AuditLog.objects.create(
                username_snapshot=request.session.get('admin_username', 'admin'),
                user_role_snapshot='admin',
                category='admin',
                action=f'Created supervisor: {username}',
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            messages.success(request, _(f'Supervisor {username} created'))

        elif action == 'create_exam_staff':
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '').strip()

            if not username or not password:
                messages.error(request, _('Username and password are required'))
                return redirect('admin_panel:supervisors')

            if CustomUser.objects.filter(username=username).exists():
                messages.error(request, _('Username already exists'))
                return redirect('admin_panel:supervisors')

            user = CustomUser.objects.create_user(
                username=username,
                password=password,
                role=UserRole.EXAM_STAFF,
                is_staff=False,
                is_superuser=False
            )
            AuditLog.objects.create(
                username_snapshot=request.session.get('admin_username', 'admin'),
                user_role_snapshot='admin',
                category='admin',
                action=f'Created exam staff user: {username}',
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            messages.success(request, _(f'Exam Staff user "{username}" created successfully.'))

        elif action == 'delete_exam_staff':
            user_id = request.POST.get('user_id')
            try:
                user = CustomUser.objects.get(id=user_id, role=UserRole.EXAM_STAFF)
                username = user.username
                user.delete()
                messages.success(request, _(f'Exam Staff user "{username}" deleted successfully.'))
            except CustomUser.DoesNotExist:
                messages.error(request, _('Exam Staff user not found.'))

        elif action == 'deactivate':
            sup_id = request.POST.get('supervisor_id')
            try:
                sup = SupervisorAccount.objects.get(id=sup_id)
                sup.is_active = False
                sup.user.is_active = False
                sup.save()
                sup.user.save()
                messages.success(request, _('Supervisor deactivated'))
            except SupervisorAccount.DoesNotExist:
                messages.error(request, _('Supervisor not found'))

        elif action == 'delete':
            sup_id = request.POST.get('supervisor_id')
            try:
                sup = SupervisorAccount.objects.get(id=sup_id)
                username = sup.user.username
                sup.user.delete()
                messages.success(request, _(f'Supervisor {username} deleted'))
            except SupervisorAccount.DoesNotExist:
                messages.error(request, _('Supervisor not found'))

        elif action == 'clear_logs':
            deleted_count, _deleted_dict = VerificationLog.objects.filter(verification_type='exam_day').delete()
            AuditLog.objects.create(
                username_snapshot=request.session.get('admin_username', 'admin'),
                user_role_snapshot='admin',
                category='data',
                action=f'Cleared supervisor activity logs: {deleted_count} deleted',
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            messages.success(request, _(f'Cleared {deleted_count} exam day activity logs.'))

        elif action == 'import_supervisors':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, _('Please upload an Excel file'))
                return redirect('admin_panel:supervisors')
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active
                
                headers = [str(sheet.cell(1, c).value).strip().lower() if sheet.cell(1, c).value else '' for c in range(1, sheet.max_column + 1)]
                
                def find_col(names):
                    for idx, h in enumerate(headers):
                        if any(n in h for n in names):
                            return idx + 1
                    return None

                col_name = find_col(['full name', 'fullname', 'name', 'ism', 'familya', 'to\'liq ism'])
                col_user = find_col(['username', 'login', 'user', 'foydalanuvchi'])
                col_pass = find_col(['password', 'parol', 'pass'])

                if not col_name: col_name = 1
                if not col_user: col_user = 2
                if not col_pass: col_pass = 3

                imported_count = 0
                errors = []
                for r in range(2, sheet.max_row + 1):
                    name_val = str(sheet.cell(r, col_name).value or '').strip()
                    user_val = str(sheet.cell(r, col_user).value or '').strip().lower()
                    pass_val = str(sheet.cell(r, col_pass).value or '').strip()

                    if not name_val and not user_val:
                        continue
                    if not user_val or not pass_val or not name_val:
                        errors.append(f"Row {r}: Username, password and full name are required")
                        continue

                    if CustomUser.objects.filter(username=user_val).exists():
                        errors.append(f"Row {r}: Username '{user_val}' already exists")
                        continue

                    user = CustomUser.objects.create_supervisor(username=user_val, password=pass_val)
                    SupervisorAccount.objects.create(user=user, full_name=name_val)
                    imported_count += 1

                AuditLog.objects.create(
                    username_snapshot=request.session.get('admin_username', 'admin'),
                    user_role_snapshot='admin',
                    category='data',
                    action=f'Imported {imported_count} supervisors via Excel',
                    ip_address=request.META.get('REMOTE_ADDR'),
                )
                if imported_count > 0:
                    messages.success(request, _(f'Successfully imported {imported_count} supervisors!'))
                if errors:
                    messages.warning(request, _(f'Warnings: {", ".join(errors[:5])}'))
            except Exception as e:
                messages.error(request, _(f'Failed to parse Excel: {str(e)}'))

        return redirect('admin_panel:supervisors')


@admin_required_class
class SupervisorLogsExportExcelView(View):
    def get(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        logs = VerificationLog.objects.filter(
            verification_type='exam_day',
            supervisor__isnull=False
        ).select_related('applicant_profile', 'supervisor').order_by('-verified_at')

        wb = Workbook()
        ws = wb.active
        ws.title = 'Supervisor Activity'
        
        headers = ['Timestamp', 'Supervisor', 'Candidate Full Name', 'Admission ID', 'Passport Number', 'Result', 'Score', 'Details']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        for log in logs:
            sup_name = ""
            if log.supervisor:
                if hasattr(log.supervisor, 'supervisor_account') and log.supervisor.supervisor_account:
                    sup_name = log.supervisor.supervisor_account.full_name
                else:
                    sup_name = log.supervisor.username
            
            ws.append([
                log.verified_at.strftime('%Y-%m-%d %H:%M:%S') if log.verified_at else '',
                sup_name,
                log.applicant_profile.full_name if log.applicant_profile else '',
                log.applicant_profile.admission_id if log.applicant_profile else '',
                log.applicant_profile.passport_number if log.applicant_profile else '',
                log.get_result_display() or '',
                f'{log.score:.1f}%' if log.score else '',
                log.notes or 'Face match verified',
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="supervisor_activity_logs.xlsx"'
        return response


# ─── QR MANAGEMENT ───────────────────────────────────────────────────────────

@admin_required_class
class QRManagementView(View):
    template_name = 'admin_panel/qr_management.html'

    def get(self, request):
        query = request.GET.get('q', '').strip()
        qr_codes = QRCode.objects.select_related('applicant_profile').order_by('-generated_at')
        if query:
            qr_codes = qr_codes.filter(
                Q(token__icontains=query) |
                Q(applicant_profile__first_name__icontains=query) |
                Q(applicant_profile__last_name__icontains=query) |
                Q(applicant_profile__passport_number__icontains=query) |
                Q(applicant_profile__admission_id__icontains=query)
            )
        paginator = Paginator(qr_codes, 25)
        page = paginator.get_page(request.GET.get('page'))
        return render(request, self.template_name, {
            'page_title': _('QR Management'),
            'qr_codes': page,
            'query': query,
        })

    def post(self, request):
        action = request.POST.get('action')
        qr_id = request.POST.get('qr_id')
        if action == 'revoke' and qr_id:
            try:
                qr = QRCode.objects.get(id=qr_id)
                qr.status = 'revoked'
                qr.save()
                messages.success(request, _('QR code revoked'))
            except QRCode.DoesNotExist:
                messages.error(request, _('QR code not found'))
        elif action == 'regenerate' and qr_id:
            try:
                qr = QRCode.objects.get(id=qr_id)
                qr.delete()
                from apps.qr_module.generator import generate_qr_code
                generate_qr_code(qr.applicant_profile)
                messages.success(request, _('QR code regenerated'))
            except Exception as e:
                messages.error(request, str(e))
        return redirect('admin_panel:qr-management')


# ─── AUDIT LOGS ──────────────────────────────────────────────────────────────

@admin_required_class
class AuditLogView(View):
    template_name = 'admin_panel/audit_logs.html'

    def get(self, request):
        logs = AuditLog.objects.order_by('-timestamp')
        category = request.GET.get('category', '')
        query = request.GET.get('q', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        if category:
            logs = logs.filter(category=category)
        if query:
            logs = logs.filter(
                Q(action__icontains=query) |
                Q(username_snapshot__icontains=query) |
                Q(ip_address__icontains=query)
            )
        if date_from:
            logs = logs.filter(timestamp__date__gte=date_from)
        if date_to:
            logs = logs.filter(timestamp__date__lte=date_to)

        paginator = Paginator(logs, 50)
        page = paginator.get_page(request.GET.get('page'))

        return render(request, self.template_name, {
            'page_title': _('Audit Logs'),
            'logs': page,
            'categories': AuditLog.ACTION_CATEGORIES,
            'selected_category': category,
            'query': query,
            'date_from': date_from,
            'date_to': date_to,
        })

    def post(self, request):
        action = request.POST.get('action')
        
        if action == 'clear_logs':
            deleted_count, _deleted_dict = AuditLog.objects.all().delete()
            AuditLog.objects.create(
                username_snapshot=request.session.get('admin_username', 'admin'),
                user_role_snapshot='admin',
                category='security',
                action='Cleared system audit logs',
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            messages.success(request, _(f'Cleared {deleted_count} system audit logs.'))
            
        elif action == 'import_logs':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, _('Please upload an Excel file'))
                return redirect('admin_panel:audit-logs')
            try:
                import openpyxl
                from django.utils.dateparse import parse_datetime
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = wb.active
                
                headers = [str(sheet.cell(1, c).value).strip().lower() if sheet.cell(1, c).value else '' for c in range(1, sheet.max_column + 1)]
                
                def find_col(names):
                    for idx, h in enumerate(headers):
                        if any(n in h for n in names):
                            return idx + 1
                    return None

                col_time = find_col(['timestamp', 'time', 'date'])
                col_user = find_col(['user', 'username', 'operator'])
                col_role = find_col(['role', 'level'])
                col_cat = find_col(['category', 'type'])
                col_act = find_col(['action', 'details', 'operation'])
                col_ip = find_col(['ip address', 'ip_address', 'ip'])
                col_success = find_col(['status', 'success'])

                if not col_time: col_time = 1
                if not col_user: col_user = 2
                if not col_act: col_act = 3

                imported_count = 0
                for r in range(2, sheet.max_row + 1):
                    time_str = str(sheet.cell(r, col_time).value or '').strip()
                    user_val = str(sheet.cell(r, col_user).value or '').strip()
                    act_val = str(sheet.cell(r, col_act).value or '').strip()
                    role_val = str(sheet.cell(r, col_role).value or '').strip() if col_role else 'admin'
                    cat_val = str(sheet.cell(r, col_cat).value or '').strip() if col_cat else 'system'
                    ip_val = str(sheet.cell(r, col_ip).value or '').strip() if col_ip else '127.0.0.1'
                    status_val = str(sheet.cell(r, col_success).value or '').strip().lower() if col_success else 'success'

                    if not act_val:
                        continue

                    timestamp = None
                    if time_str:
                        try:
                            timestamp = parse_datetime(time_str)
                        except Exception:
                            pass
                    if not timestamp:
                        timestamp = timezone.now()

                    success_flag = status_val in ('success', 'true', '1', 'yes')

                    AuditLog.objects.create(
                        timestamp=timestamp,
                        username_snapshot=user_val,
                        user_role_snapshot=role_val,
                        category=cat_val,
                        action=act_val,
                        ip_address=ip_val,
                        success=success_flag,
                    )
                    imported_count += 1

                AuditLog.objects.create(
                    username_snapshot=request.session.get('admin_username', 'admin'),
                    user_role_snapshot='admin',
                    category='data',
                    action=f'Imported {imported_count} audit logs via Excel',
                    ip_address=request.META.get('REMOTE_ADDR'),
                )
                messages.success(request, _(f'Successfully imported {imported_count} audit logs!'))
            except Exception as e:
                messages.error(request, _(f'Failed to parse Excel: {str(e)}'))

        return redirect('admin_panel:audit-logs')


@admin_required_class
class AuditLogsExportExcelView(View):
    def get(self, request):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        logs = AuditLog.objects.order_by('-timestamp')
        category = request.GET.get('category', '')
        query = request.GET.get('q', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        if category:
            logs = logs.filter(category=category)
        if query:
            logs = logs.filter(
                Q(action__icontains=query) |
                Q(username_snapshot__icontains=query) |
                Q(ip_address__icontains=query)
            )
        if date_from:
            logs = logs.filter(timestamp__date__gte=date_from)
        if date_to:
            logs = logs.filter(timestamp__date__lte=date_to)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Audit Logs'
        
        headers = ['Timestamp', 'Username', 'Role', 'Category', 'Action', 'Status', 'IP Address']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F4E78')
            cell.alignment = Alignment(horizontal='center')

        for log in logs:
            ws.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
                log.username_snapshot or 'System',
                log.user_role_snapshot or '',
                log.category or '',
                log.action or '',
                'Success' if log.success else 'Failed',
                log.ip_address or '',
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="system_audit_logs.xlsx"'
        return response





# ─── SYSTEM SETTINGS ─────────────────────────────────────────────────────────

@admin_required_class
class SystemSettingsView(View):
    template_name = 'admin_panel/settings.html'

    def get(self, request):
        from apps.accounts.models import SystemSetting, ExamVenueConfig
        # Ensure setting exists (singleton pattern)
        setting = SystemSetting.objects.first()
        if not setting:
            setting = SystemSetting.objects.create(qr_domain='id.akhu.uz')
            
        venues = ExamVenueConfig.objects.all().order_by('region')
        
        return render(request, self.template_name, {
            'page_title': _('System Settings'),
            'setting': setting,
            'venues': venues,
        })

    def post(self, request):
        from apps.accounts.models import SystemSetting, ExamVenueConfig, ApplicantProfile
        from django.db import transaction
        from django.conf import settings
        import time

        action = request.POST.get('action', 'save_settings')
        start_time = time.time()

        if action == 'add_region':
            region_name = request.POST.get('region_name', '').strip()
            venue_name = request.POST.get('venue_name', '').strip()
            exam_date_str = request.POST.get('exam_date', '').strip()
            arrival_time_str = request.POST.get('arrival_time', '').strip()
            location_link = request.POST.get('location_link', '').strip()

            if not region_name:
                messages.error(request, _('Region name is required.'))
                return redirect('admin_panel:settings')

            if ExamVenueConfig.objects.filter(region=region_name).exists():
                messages.error(request, _('A region with name "{name}" already exists.').format(name=region_name))
                return redirect('admin_panel:settings')

            try:
                ExamVenueConfig.objects.create(
                    region=region_name,
                    venue_name=venue_name,
                    exam_date=exam_date_str,
                    arrival_time=arrival_time_str,
                    location_link=location_link
                )
                messages.success(request, _('Region "{name}" created successfully.').format(name=region_name))
            except Exception as e:
                messages.error(request, _('Failed to create region: {error}').format(error=str(e)))

            return redirect('admin_panel:settings')

        elif action == 'delete_region':
            venue_id = request.POST.get('venue_id')
            try:
                venue = ExamVenueConfig.objects.get(id=venue_id)
                region_name = venue.region
                venue.delete()
                messages.success(request, _('Region "{name}" deleted successfully.').format(name=region_name))
            except ExamVenueConfig.DoesNotExist:
                messages.error(request, _('Region not found.'))
            except Exception as e:
                messages.error(request, _('Failed to delete region: {error}').format(error=str(e)))

            return redirect('admin_panel:settings')

        elif action == 'save_settings':
            logger.info("Permit Release: Settings save initiated by admin.")
            try:
                with transaction.atomic():
                    # 1. Save general settings (singleton pattern)
                    setting = SystemSetting.objects.first()
                    if not setting:
                        setting = SystemSetting.objects.create(qr_domain='id.akhu.uz')
                    setting.qr_domain = request.POST.get('qr_domain', 'id.akhu.uz').strip()
                    setting.permits_released = 'permits_released' in request.POST
                    
                    release_date_str = request.POST.get('permit_release_date', '').strip()
                    if release_date_str:
                        try:
                            from datetime import datetime
                            setting.permit_release_date = datetime.strptime(release_date_str, '%Y-%m-%d').date()
                        except ValueError:
                            setting.permit_release_date = None
                    else:
                        setting.permit_release_date = None
                    setting.save()
                    
                    # 2. Save venue configs
                    venues = ExamVenueConfig.objects.all()
                    for venue in venues:
                        old_region = venue.region
                        new_region = request.POST.get(f'region_name_{venue.id}', '').strip()
                        venue_name = request.POST.get(f'venue_{venue.id}', '').strip()
                        exam_date_str = request.POST.get(f'date_{venue.id}', '').strip()
                        arrival_time_str = request.POST.get(f'arrival_{venue.id}', '').strip()
                        location_link = request.POST.get(f'location_link_{venue.id}', '').strip()
                        
                        if new_region and new_region != old_region:
                            if ExamVenueConfig.objects.exclude(id=venue.id).filter(region=new_region).exists():
                                raise ValueError(_('Region name "{name}" already exists.').format(name=new_region))
                            
                            venue.region = new_region
                            # Propagate rename to all applicant profiles
                            ApplicantProfile.objects.filter(selected_region=old_region).update(selected_region=new_region)

                        venue.venue_name = venue_name
                        venue.location_link = location_link
                        venue.exam_date = exam_date_str
                        venue.arrival_time = arrival_time_str
                        venue.save()
                    
                    # 3. Synchronize verified candidates
                    verified_profiles = list(ApplicantProfile.objects.filter(
                        user__verification_sessions__face_profile__status='verified'
                    ).select_related('user').prefetch_related(
                        'user__verification_sessions__face_profile', 
                        'qr_code'
                    ).distinct())

                    logger.info(f"Permit Release: found {len(verified_profiles)} verified candidate profiles to sync.")

                    venue_configs = {config.region: config for config in ExamVenueConfig.objects.all()}
                    
                    updated_profiles = []
                    for profile in verified_profiles:
                        changed = False
                        if setting.permits_released:
                            if profile.selected_region in venue_configs:
                                conf = venue_configs[profile.selected_region]
                                if profile.exam_venue != conf.venue_name:
                                    profile.exam_venue = conf.venue_name
                                    changed = True
                                if profile.exam_date != conf.exam_date:
                                    profile.exam_date = conf.exam_date
                                    changed = True
                                if profile.arrival_time != conf.arrival_time:
                                    profile.arrival_time = conf.arrival_time
                                    changed = True
                        else:
                            if profile.exam_venue != "":
                                profile.exam_venue = ""
                                changed = True
                            if profile.exam_date != "":
                                profile.exam_date = ""
                                changed = True
                            if profile.arrival_time != "":
                                profile.arrival_time = ""
                                changed = True
                                
                        if changed:
                            updated_profiles.append(profile)

                    if updated_profiles:
                        ApplicantProfile.objects.bulk_update(updated_profiles, ['exam_venue', 'exam_date', 'arrival_time'])
                        
                    logger.info(f"Permit Release: synchronized {len(updated_profiles)} candidate profiles in database.")

                # Process QR generation post-commit
                qr_generated_count = 0
                if setting.permits_released:
                    from apps.qr_module.services import generate_applicant_qr
                    for profile in verified_profiles:
                        has_qr = False
                        try:
                            has_qr = bool(profile.qr_code)
                        except Exception:
                            pass
                            
                        if not has_qr:
                            try:
                                generate_applicant_qr(profile)
                                qr_generated_count += 1
                            except Exception as e:
                                logger.error(f"Failed to generate QR code for verified applicant {profile.admission_id}: {e}")

                duration = time.time() - start_time
                logger.info(f"Permit Release: process completed. Synced profiles: {len(updated_profiles)}, Generated QRs: {qr_generated_count}. Duration: {duration:.2f}s.")

                messages.success(request, _('System settings and exam venues updated successfully.'))
                if qr_generated_count > 0:
                    messages.info(request, _("Generated QR codes for {count} verified candidates.").format(count=qr_generated_count))
                
                return redirect('admin_panel:settings')

            except Exception as e:
                duration = time.time() - start_time
                logger.error(f"Permit Release: process failed after {duration:.2f}s, transaction rolled back. Error: {e}")
                messages.error(request, _("An error occurred: {error}").format(error=str(e)))
                return redirect('admin_panel:settings')


# ─── REPORTS ─────────────────────────────────────────────────────────────────

@admin_required_class
class ReportsView(TemplateView):
    template_name = 'admin_panel/reports.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = _('Reports')
        return ctx


# ─── STATISTICS ──────────────────────────────────────────────────────────────

@admin_required_class
class StatisticsView(TemplateView):
    template_name = 'admin_panel/statistics.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # Total registered applicants
        total_applicants = ApplicantProfile.objects.count()

        # Verified applicants (locked profile or verified status)
        verified_profiles = ApplicantProfile.objects.filter(
            Q(is_locked=True) | Q(user__verification_sessions__face_profile__status='verified')
        ).distinct()
        verified_count = verified_profiles.count()

        pending_count = total_applicants - verified_count
        if pending_count < 0:
            pending_count = 0

        pass_rate = round((verified_count / total_applicants * 100), 1) if total_applicants > 0 else 0.0

        # Regional Breakdown aggregation
        raw_regions = ApplicantProfile.objects.values('selected_region').annotate(
            total=Count('id', distinct=True),
            verified=Count('id', filter=Q(is_locked=True) | Q(user__verification_sessions__face_profile__status='verified'), distinct=True)
        ).order_by('-total')

        configured_venues = list(ExamVenueConfig.objects.values_list('region', flat=True))
        
        region_map = {}
        for item in raw_regions:
            reg_raw = (item['selected_region'] or '').strip()
            reg = reg_raw if reg_raw else _('Unassigned').strip()
            if reg not in region_map:
                region_map[reg] = {'total': 0, 'verified': 0}
            region_map[reg]['total'] += item['total']
            region_map[reg]['verified'] += item['verified']

        for conf_region in configured_venues:
            if conf_region:
                cr = conf_region.strip()
                if cr and cr not in region_map:
                    region_map[cr] = {'total': 0, 'verified': 0}

        region_stats = []
        for reg_name, counts in region_map.items():
            tot = counts['total']
            ver = counts['verified']
            pend = tot - ver if tot >= ver else 0
            pct = round((ver / tot * 100), 1) if tot > 0 else 0.0
            region_stats.append({
                'region': reg_name,
                'total': tot,
                'verified': ver,
                'pending': pend,
                'pass_rate': pct,
            })

        region_stats.sort(key=lambda x: (x['total'], x['verified']), reverse=True)

        # Program Breakdown
        raw_programs = ApplicantProfile.objects.values('program').annotate(
            total=Count('id', distinct=True),
            verified=Count('id', filter=Q(is_locked=True) | Q(user__verification_sessions__face_profile__status='verified'), distinct=True)
        ).order_by('-total')

        program_stats = []
        for item in raw_programs:
            prog_raw = (item['program'] or '').strip()
            prog = prog_raw if prog_raw else _('General')
            tot = item['total']
            ver = item['verified']
            pend = tot - ver if tot >= ver else 0
            pct = round((ver / tot * 100), 1) if tot > 0 else 0.0
            program_stats.append({
                'program': prog,
                'total': tot,
                'verified': ver,
                'pending': pend,
                'pass_rate': pct,
            })

        # Daily Trends for Chart (Last 7 Days)
        from datetime import timedelta
        today = timezone.now().date()
        daily_trends = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            cnt = VerificationLog.objects.filter(verified_at__date=d, result='verified').count()
            daily_trends.append({
                'date': d.strftime('%d.%m'),
                'full_date': str(d),
                'count': cnt,
            })

        ctx.update({
            'page_title': _('Verification Statistics'),
            'active': 'statistics',
            'total_applicants': total_applicants,
            'verified_count': verified_count,
            'pending_count': pending_count,
            'pass_rate': pass_rate,
            'region_stats': region_stats,
            'program_stats': program_stats,
            'daily_trends': daily_trends,
        })
        return ctx


@admin_required_class
class StatisticsExportExcelView(View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Regional Statistics"
        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="0A1628", end_color="0A1628", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="0A1628")

        ws.merge_cells("A1:F1")
        ws["A1"] = "AKHU Verification Platform — Regional Statistics"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(size=10, italic=True, color="555555")

        headers = ["#", "Region / Location", "Total Applicants", "Face ID Verified", "Pending Verification", "Pass Rate (%)"]
        ws.append([])
        ws.append(headers)

        ws.row_dimensions[4].height = 26
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        raw_regions = ApplicantProfile.objects.values('selected_region').annotate(
            total=Count('id', distinct=True),
            verified=Count('id', filter=Q(is_locked=True) | Q(user__verification_sessions__face_profile__status='verified'), distinct=True)
        ).order_by('-total')

        configured_venues = list(ExamVenueConfig.objects.values_list('region', flat=True))
        region_map = {}
        for item in raw_regions:
            reg_raw = (item['selected_region'] or '').strip()
            reg = reg_raw if reg_raw else "Unassigned"
            if reg not in region_map:
                region_map[reg] = {'total': 0, 'verified': 0}
            region_map[reg]['total'] += item['total']
            region_map[reg]['verified'] += item['verified']

        for conf_region in configured_venues:
            if conf_region:
                cr = conf_region.strip()
                if cr and cr not in region_map:
                    region_map[cr] = {'total': 0, 'verified': 0}

        region_stats = []
        for reg_name, counts in region_map.items():
            tot = counts['total']
            ver = counts['verified']
            pend = tot - ver if tot >= ver else 0
            pct = round((ver / tot * 100), 1) if tot > 0 else 0.0
            region_stats.append({
                'region': reg_name,
                'total': tot,
                'verified': ver,
                'pending': pend,
                'pass_rate': pct,
            })
        region_stats.sort(key=lambda x: (x['total'], x['verified']), reverse=True)

        row_idx = 5
        sum_total = 0
        sum_verified = 0
        sum_pending = 0

        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        for idx, item in enumerate(region_stats, 1):
            sum_total += item['total']
            sum_verified += item['verified']
            sum_pending += item['pending']

            ws.append([
                idx,
                item['region'],
                item['total'],
                item['verified'],
                item['pending'],
                f"{item['pass_rate']}%"
            ])
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, 7):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border
                c.alignment = Alignment(horizontal="center" if col_idx in [1, 3, 4, 5, 6] else "left", vertical="center")
            row_idx += 1

        # Summary Total Row
        total_pct = round((sum_verified / sum_total * 100), 1) if sum_total > 0 else 0.0
        ws.append([
            "",
            "TOTAL / JAMI",
            sum_total,
            sum_verified,
            sum_pending,
            f"{total_pct}%"
        ])
        ws.row_dimensions[row_idx].height = 24
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        total_font = Font(name="Calibri", size=11, bold=True, color="0A1628")
        for col_idx in range(1, 7):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = total_fill
            c.font = total_font
            c.border = Border(top=Side(style='medium', color='0A1628'), bottom=Side(style='double', color='0A1628'))
            c.alignment = Alignment(horizontal="center" if col_idx in [1, 3, 4, 5, 6] else "left", vertical="center")

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="AKHU_Regional_Statistics_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response


# ─── LIVENESS BIOMETRICS AUDIT ───────────────────────────────────────────────

@admin_required_class
class LivenessBiometricsView(View):
    template_name = 'admin_panel/liveness_biometrics.html'

    def get(self, request):
        query = request.GET.get('q', '').strip()

        applicants = ApplicantProfile.objects.filter(
            user__verification_sessions__face_profile__status='verified'
        ).select_related('user').prefetch_related(
            'user__verification_sessions__face_profile'
        ).distinct().order_by('-created_at')

        if query:
            applicants = applicants.filter(
                Q(applicant_id__icontains=query) |
                Q(passport_number__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(program__icontains=query)
            )

        paginator = Paginator(applicants, 10)
        page = paginator.get_page(request.GET.get('page'))

        liveness_data = []
        for app in page:
            liveness_data.append({
                'profile': app,
                'face_profile': _verified_liveness_face_profile(app),
            })

        return render(request, self.template_name, {
            'page_title': _('Liveness Biometrics'),
            'applicants': page,
            'liveness_data': liveness_data,
            'query': query,
            'active': 'liveness',
        })


def _safe_zip_folder_name(value: str, fallback: str) -> str:
    import re

    raw = (value or fallback or 'applicant').strip()
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', '_', raw)
    cleaned = re.sub(r'\s+', '_', cleaned).strip('._ ')
    return cleaned[:120] or 'applicant'


def _verified_liveness_face_profile(app_profile):
    return FaceProfile.objects.filter(
        session__user=app_profile.user,
        status='verified',
    ).select_related('session').order_by('-created_at').first()


def _write_liveness_zip_folder(zip_file, app_profile, face_profile, used_folders=None) -> int:
    import os

    fallback = app_profile.admission_id or app_profile.passport_number or str(app_profile.id)
    folder = _safe_zip_folder_name(app_profile.applicant_id, fallback)

    if used_folders is not None:
        base_folder = folder
        counter = 2
        while folder in used_folders:
            folder = f'{base_folder}_{counter}'
            counter += 1
        used_folders.add(folder)

    photos = {
        '01_frontal.jpg': face_profile.selfie_image,
        '02_left.jpg': face_profile.selfie_left,
        '03_right.jpg': face_profile.selfie_right,
        '04_up.jpg': face_profile.selfie_up,
    }

    written = 0
    info_lines = [
        f'Applicant ID: {app_profile.applicant_id or "-"}',
        f'Admission ID: {app_profile.admission_id or "-"}',
        f'Full name: {app_profile.full_name}',
        f'Passport: {app_profile.passport_number}',
        '',
        'Files:',
    ]

    for filename, img_field in photos.items():
        if not img_field:
            info_lines.append(f'- {filename}: missing')
            continue
        try:
            path = img_field.path
            if os.path.exists(path):
                with open(path, 'rb') as image_file:
                    zip_file.writestr(f'{folder}/{filename}', image_file.read())
                written += 1
                info_lines.append(f'- {filename}: included')
            else:
                info_lines.append(f'- {filename}: file not found')
        except Exception as exc:
            logger.warning('Could not add liveness image for %s: %s', app_profile.id, exc)
            info_lines.append(f'- {filename}: unavailable')

    zip_file.writestr(f'{folder}/README.txt', '\n'.join(info_lines))
    return written


@admin_required_class
class DownloadLivenessZipView(View):
    def get(self, request, profile_id):
        import io
        import zipfile

        app_profile = get_object_or_404(ApplicantProfile, id=profile_id)
        face_profile = _verified_liveness_face_profile(app_profile)
        if not face_profile:
            messages.error(request, _('No verified liveness data found for this applicant.'))
            return redirect('admin_panel:liveness-biometrics')

        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            _write_liveness_zip_folder(zip_file, app_profile, face_profile)

        zip_buffer.seek(0)

        AuditLog.objects.create(
            username_snapshot=request.session.get('admin_username', 'admin'),
            user_role_snapshot='admin',
            category='data',
            action=f'Downloaded liveness ZIP for {app_profile.applicant_id or app_profile.passport_number}',
            ip_address=request.META.get('REMOTE_ADDR'),
            success=True,
        )

        zip_name = f"{_safe_zip_folder_name(app_profile.applicant_id, app_profile.passport_number)}_liveness.zip"
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
        zip_buffer.close()
        return response


@admin_required_class
class DownloadAllLivenessZipView(View):
    def get(self, request):
        import io
        import zipfile

        applicants = ApplicantProfile.objects.filter(
            user__verification_sessions__face_profile__status='verified'
        ).select_related('user').prefetch_related(
            'user__verification_sessions__face_profile'
        ).distinct()

        if not applicants.exists():
            messages.warning(request, _('No verified applicants found with liveness data.'))
            return redirect('admin_panel:liveness-biometrics')

        zip_buffer = io.BytesIO()
        used_folders = set()
        included_applicants = 0
        included_images = 0
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for app in applicants:
                face_profile = _verified_liveness_face_profile(app)
                if not face_profile:
                    continue

                included_images += _write_liveness_zip_folder(zip_file, app, face_profile, used_folders)
                included_applicants += 1

            zip_file.writestr(
                'MANIFEST.txt',
                f'Total applicants: {included_applicants}\nTotal images: {included_images}\nFolder naming: Applicant ID, with fallback to Admission ID or Passport.\n'
            )

        zip_buffer.seek(0)

        AuditLog.objects.create(
            username_snapshot=request.session.get('admin_username', 'admin'),
            user_role_snapshot='admin',
            category='data',
            action=f'Bulk downloaded liveness ZIP for {included_applicants} applicants',
            ip_address=request.META.get('REMOTE_ADDR'),
            success=True,
        )

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="all_applicants_liveness.zip"'
        zip_buffer.close()
        return response
