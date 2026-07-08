"""
excel_merge/numbering.py

Responsible for:
- Detecting if there is a "T/r" (Tartib raqami) column in the headers
- If present, sequentially renumbering the rows (1, 2, 3, ...) in that column for all data rows
"""
import logging
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def renumber_tr_column(ws: Worksheet, header_size: int) -> bool:
    """
    Search the header rows for a column named exactly "T/r" (case-insensitive, stripped).
    If found, renumber all data rows starting from 1.
    
    Returns True if a "T/r" column was found and renumbered, False otherwise.
    """
    tr_col_idx = None

    # Scan header rows to find the "T/r" column
    for r_idx in range(1, header_size + 1):
        for c_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=r_idx, column=c_idx).value
            if isinstance(cell_val, str) and cell_val.strip().lower() == "t/r":
                tr_col_idx = c_idx
                break
        if tr_col_idx is not None:
            break

    if tr_col_idx is None:
        logger.info("No 'T/r' column detected in the header rows.")
        return False

    logger.info("Found 'T/r' column at index %d. Renumbering data rows...", tr_col_idx)

    seq_num = 1
    for r_idx in range(header_size + 1, ws.max_row + 1):
        # Renumber only if the row has some data or values, to avoid writing to empty/blank spacer rows
        # But if the row contains values, we always renumber it.
        row_has_data = any(ws.cell(row=r_idx, column=c).value is not None for c in range(1, ws.max_column + 1) if c != tr_col_idx and c != ws.max_column)
        
        if row_has_data:
            ws.cell(row=r_idx, column=tr_col_idx).value = seq_num
            seq_num += 1

    logger.info("Successfully renumbered %d data rows.", seq_num - 1)
    return True
