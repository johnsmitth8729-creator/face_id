"""
excel_merge/merger.py

Responsible for:
- Unmerging cell ranges and propagating top-left cell values & styles to all cells in the range
- Determining the number of header rows by finding the longest common prefix of identical rows across all workbooks
- Executing the row-by-row merging logic while copying cell styles and row heights
- Appending the "Source File" column to the end of each copied row
"""
import logging
from typing import List, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from excel_merge.formatter import copy_cell_style, copy_row_formatting, update_column_widths
from excel_merge.loader import is_row_hidden

logger = logging.getLogger(__name__)


def unmerge_and_propagate(ws: Worksheet):
    """
    Find all merged cell ranges in ws. For each range:
    1. Retrieve the value and style of the top-left cell.
    2. Unmerge the range.
    3. Copy the value and style to every cell in the range.
    """
    # Create a list of ranges to avoid modifying the collection during iteration
    merged_ranges = list(ws.merged_cells.ranges)
    if not merged_ranges:
        return

    logger.info("Unmerging %d ranges in sheet '%s'...", len(merged_ranges), ws.title)

    for rng in merged_ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        top_left_cell = ws.cell(row=min_row, column=min_col)
        val = top_left_cell.value

        # Unmerge the range
        ws.unmerge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col
        )

        # Propagate the top-left cell value and styling to all cells in the range
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = val
                copy_cell_style(top_left_cell, cell)


def determine_header_rows(loaded_sheets: List[Tuple[str, openpyxl.Workbook, Worksheet]]) -> int:
    """
    Identify the number of header rows by finding the longest common prefix of rows
    that are identical in values across all workbooks.
    """
    if len(loaded_sheets) <= 1:
        # If there's only 1 workbook, there are no duplicate headers to deduplicate.
        # We can scan the first few rows to look for a header. We default to 1.
        return 1

    # Read the values of the first 100 rows for each worksheet (skipping hidden rows)
    all_sheet_rows = []
    for filename, wb, ws in loaded_sheets:
        rows_vals = []
        for r_idx in range(1, min(ws.max_row + 1, 100)):
            if is_row_hidden(ws, r_idx):
                continue
            # Get cell values as a tuple of strings/values
            row_vals = tuple(cell.value for cell in ws[r_idx])
            rows_vals.append(row_vals)
        all_sheet_rows.append(rows_vals)

    # Compare row-by-row across all loaded sheets
    first_sheet_rows = all_sheet_rows[0]
    header_rows_count = 0
    max_compare = min(len(rows) for rows in all_sheet_rows)

    for r_idx in range(max_compare):
        row_first = first_sheet_rows[r_idx]
        match_all = True
        for sheet_idx in range(1, len(all_sheet_rows)):
            if all_sheet_rows[sheet_idx][r_idx] != row_first:
                match_all = False
                break
        if match_all:
            header_rows_count += 1
        else:
            break

    logger.info("Detected %d duplicate header row(s) based on prefix matching.", header_rows_count)
    return header_rows_count


def merge_worksheets(
    loaded_sheets: List[Tuple[str, openpyxl.Workbook, Worksheet]]
) -> openpyxl.Workbook:
    """
    Merges all loaded worksheets into a single workbook.
    - Preserves styling, column widths, row heights.
    - Deduplicates headers.
    - Unmerges cells and propagates values.
    - Appends the 'Source File' column.
    """
    # Create the output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Merged Data"

    # Pre-process all source sheets to unmerge cells and propagate values
    for filename, wb, ws in loaded_sheets:
        unmerge_and_propagate(ws)

    # Determine header size
    header_size = determine_header_rows(loaded_sheets)

    dest_row_idx = 1
    master_col_count = 0

    # 1. Process the first workbook (keeps headers + data)
    first_filename, _, first_ws = loaded_sheets[0]
    update_column_widths(first_ws, out_ws, is_first=True)
    master_col_count = first_ws.max_column

    logger.info("Merging first workbook '%s' (headers and data)...", first_filename)

    for r_idx in range(1, first_ws.max_row + 1):
        if is_row_hidden(first_ws, r_idx):
            continue

        # Copy row cells
        for c_idx in range(1, master_col_count + 1):
            src_cell = first_ws.cell(row=r_idx, column=c_idx)
            dest_cell = out_ws.cell(row=dest_row_idx, column=c_idx)
            dest_cell.value = src_cell.value
            copy_cell_style(src_cell, dest_cell)

        # Copy row dimensions
        copy_row_formatting(first_ws, r_idx, out_ws, dest_row_idx)

        # Add Source File column
        source_cell = out_ws.cell(row=dest_row_idx, column=master_col_count + 1)
        if r_idx <= header_size:
            source_cell.value = "Source File"
        else:
            source_cell.value = first_filename

        # Match formatting of the last cell in the row for the source column cell
        last_cell = first_ws.cell(row=r_idx, column=master_col_count)
        copy_cell_style(last_cell, source_cell)

        dest_row_idx += 1

    # 2. Process every subsequent workbook (skips duplicate headers)
    for filename, _, ws in loaded_sheets[1:]:
        update_column_widths(ws, out_ws)
        ws_col_count = ws.max_column
        # Track if the column counts differ, but use the master_col_count for output layout consistency
        col_count = min(master_col_count, ws_col_count)

        logger.info("Merging subsequent workbook '%s' (skipping headers)...", filename)

        # We skip the duplicate header rows
        for r_idx in range(header_size + 1, ws.max_row + 1):
            if is_row_hidden(ws, r_idx):
                continue

            for c_idx in range(1, col_count + 1):
                src_cell = ws.cell(row=r_idx, column=c_idx)
                dest_cell = out_ws.cell(row=dest_row_idx, column=c_idx)
                dest_cell.value = src_cell.value
                copy_cell_style(src_cell, dest_cell)

            copy_row_formatting(ws, r_idx, out_ws, dest_row_idx)

            # Add Source File column value
            source_cell = out_ws.cell(row=dest_row_idx, column=master_col_count + 1)
            source_cell.value = filename
            
            # Format the source column cell matching the last cell style
            last_cell = ws.cell(row=r_idx, column=col_count)
            copy_cell_style(last_cell, source_cell)

            dest_row_idx += 1

    # Adjust the width of the new "Source File" column
    source_col_letter = openpyxl.utils.get_column_letter(master_col_count + 1)
    # Set it to a reasonable width to display filenames nicely
    out_ws.column_dimensions[source_col_letter].width = 25

    return out_wb
